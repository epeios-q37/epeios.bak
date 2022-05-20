# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2020 Claude SIMON (https://q37.info/s/rmnmqd49)

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import os, sys

os.chdir(os.path.dirname(os.path.realpath(__file__)))
sys.path.append("../../atlastk")

# Note to the dev: 'openpyxl' must be imported before 'atlastk',
# as the 'load_workbook' function from 'openpyxl' is overloaded.
# Applies only in DEV context.
import openpyxl
import atlastk

import updates
import importlib


print(updates.PRICE_UPDATES)

salesItem = """
<tr id="{}">
  <td>{}</td>
  <td>{}</td>
  <td>{}</td>
  <td>{}</td>
</tr>
"""

updatesItem = """
<tr id="{}">
  <td>{}</td>
  <td>{}</td>
  <td>{}</td>
</tr>
"""

modificationsItem = """
<tr xdh:onevent="View" xdh:mark="{}">
  <td>{}</td>
  <td>{}</td>
</tr>
"""

def readingUpdates(dom,modifications=None):
  global updates
  importlib.reload(updates)
  dom.setValue("updates", "")

  updatesLayout = ""

  for produce, price in updates.PRICE_UPDATES.items():
    updatesLayout += updatesItem.format(produce,produce,price,"N/A" if modifications == None else len(modifications[produce]) if produce in modifications else 0)

  dom.inner("updates", updatesLayout)
  
def reading(dom):
  dom.setValue("output", "Initialization…")

  readingUpdates(dom)

  dom.setValue('output', 'Done')
  dom.addClass("output", "hidden")

def updateWorkbook(dom):
  dom.setValue("sales", "");

  dom.setValue('output', 'Opening workbook...')

  wb = openpyxl.load_workbook("produceSales_.xlsx", read_only=True)

  sheet = wb['Sheet']

  dom.removeClass("output", "hidden")

  dom.setValue('output', 'Reading rows...')

  limit = sheet.max_row - 1	# This takes time, so it is stored.

  index = 0
  salesLayout = ""

  modifications = {}

  for row in sheet.iter_rows(min_row=2, min_col=1,max_col=3,values_only=True):
    index += 1

    produce, cost, sold = row

    if produce in updates.PRICE_UPDATES:
      modifications.setdefault(produce, [])
      modifications[produce].append(index)

    salesLayout += salesItem.format(index+1,index+1,produce,cost,round(sold,2))

    if not ( index % 2000 ) or ( index == limit ):
      dom.end('sales', salesLayout)
      dom.scrollTo(dom.lastChild('sales'))
      dom.flush()
      dom.setValue('output', 'Reading rows {}/{}'.format(index,limit))
      salesLayout = ""

  return modifications
  
def displayModifications(dom,modifications):
  modificationsLayout = ""
  for produce, lines in modifications.items():
    for line in lines:
      modificationsLayout += modificationsItem.format(line,produce, line)

  dom.inner("modifications", modificationsLayout)


def acConnect(dom):
  dom.inner("", open("Main.html").read())
  reading(dom)

def acApply(dom):
  modifications = updateWorkbook(dom)
  displayModifications(dom,modifications)
  readingUpdates(dom,modifications)

def acView(dom,id):
  dom.scrollTo(dom.getValue(id))


callbacks = {
  "": acConnect,
  "Refresh": lambda dom: readingUpdates(dom),
  "Apply": acApply,
  "View": acView
}

atlastk.launch(callbacks, None, open("Head.html").read())
