# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2020 Claude SIMON (https://q37.info/s/rmnmqd49)

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import os, sys

os.chdir(os.path.dirname(os.path.realpath(__file__)))
sys.path.append("../../atlastk")

# Note to the dev: 'openpyxl' must be imported before 'atlastk',
# as the 'load_workbook' function from 'openpyxl' is overloaded.
# Applies only in DEV context.
import openpyxl
import atlastk
import scrambler

targetTemplate = """
<tr id="tr.{Produce}" class="unselected" xdh:mark="{Produce}">
  <td class="checkbox">
    <input id="checkbox.{Produce}" type="checkbox" xdh:onevent="CheckboxClick"/>
  </td>
  <td>{Produce}</td>
  <td>{Count}</td>
</tr>
"""

sourceTemplate = """
<tr id="tr.{Produce}" class="unselected" xdh:mark="{Produce}">
  <td class="radio">
    <input id="radio.{Produce}" type="radio" name="radio" xdh:onevent="RadioClick"/>
  </td>
  <td>{Produce}</td>
  <td>{Count}</td>
</tr>
"""

sourceLabel = ""
targetLabels = []
expanded = True
workbook = None

def load(workbook):
  sheet = workbook['Sheet']
  items = {}

  for row in sheet.iter_rows(min_row=2, min_col=1,max_col=1,values_only=True):
    produce = row[0]

    items.setdefault(produce, 0)
    items[produce] += 1

  return items, (sheet.max_row - 1) / len(items)

def targetsLayout(items,limit):
  layout = ""

  for item in sorted(items.items(), key = lambda item: item[1]):
    produce, count = item

    if count > limit:
      break;

    layout += targetTemplate.format(Produce = produce, Count = count)

  return layout

def sourcesLayout(items,limit):
  layout = ""

  for item in sorted(items.items(), key = lambda item: item[0]):
    produce, count = item
    
    if count >= limit:
      layout += sourceTemplate.format(Produce = produce, Count = count)

  return layout

def fill(dom):
  items, limit = load(workbook)

  dom.after("targets",targetsLayout(items,limit))
  dom.after("sources",sourcesLayout(items,limit))

def display(dom):
  global targetLabels, sourceLabel, expanded

  fill(dom)
  targetLabels = []
  sourceLabel = ""
  expanded = True
  dom.disableElement("HideUnselected")

def acConnect(dom):
  global workbook

  dom.inner("", open("Main.html").read())

  dom.setValue("output", "Opening workbook (may take some time)…")
  dom.removeClass("output", "hidden")

  workbook = openpyxl.load_workbook("produceSales.xlsx")

  dom.setValue("output", "Scrambling…")
  scrambler.scramble(workbook)

  dom.setValue("output", "Displaying…")
  display(dom)

  dom.setValue("output", "Done.")
  dom.addClass("output", "hidden")

def acCheckboxClick(dom,id):
  global targetLabels

  mark = dom.getMark(id)

  if mark in targetLabels:
    targetLabels.remove(mark)
  else:
    targetLabels.append(mark)

  dom.toggleClass("tr.{}".format(mark),"unselected")

def acRadioClick(dom,id):
  global sourceLabel

  if sourceLabel:
    dom.toggleClass("tr.{}".format(sourceLabel),"unselected")

  sourceLabel = dom.getMark(id)

  dom.toggleClass("tr.{}".format(sourceLabel),"unselected")

def acCollapseExpand(dom):
  global expanded

  expanded = not expanded

  if expanded:
    dom.disableElement("HideUnselected")
  else:
    dom.enableElement("HideUnselected")

def acApply(dom):
  global workbook

  sheet = workbook['Sheet']

  for row in sheet.iter_rows(min_row=2, min_col=1,max_col=1):
    if row[0].value in targetLabels:
      row[0].value = sourceLabel

  dom.inner("", open("Main.html").read())
  
  display(dom)
  
callbacks = {
  "": acConnect,
  "CheckboxClick": acCheckboxClick,
  "RadioClick": acRadioClick,
  "Jump": lambda dom: dom.scrollTo("sources"),
  "CollapseExpand": acCollapseExpand,
  "Apply": acApply
}

atlastk.launch(callbacks, None, open("Head.html").read())
