# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2020 Claude SIMON (https://q37.info/s/rmnmqd49)

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import os, sys

sys.path.append("./atlastk")
sys.path.append("../atlastk")

import openpyxl
import atlastk as Atlas

tableItem = """
<tr{}>
	<td>{}</td>
	<td>{}</td>
	<td>{}</td>
	<td>{}</td>
</tr>
"""

statesItem = """
<div data-xdh-content="{State}" data-xdh-onevent="View" style="cursor: default;">{State}</div>
"""

countiesItem = """
<tr data-xdh-onevent="View" data-xdh-content="{State}.{County}" style="cursor: default;">
	<td>{County}</td>
	<td>{Pop}</td>
	<td>{Tracts}</td>
</tr>
"""

stateInCountiesItem = """
<tr id=\"{State}\" class="state">
	<td><span style=\"font-style: oblique\">State</span>: {State}</td>
	<td>{Pop}</td>
	<td>{Tracts}</td>
</tr>
"""

countyData = {}

def reading(dom):
	global countyData

	countyData= {}

	prevCounty = ""

	dom.set_content('output', 'Opening workbook...')
	wb = openpyxl.load_workbook('censuspopdata__.xlsx',read_only=True)

	sheet = wb['Population by Census Tract']

	dom.set_content('output', 'Reading rows...')

	limit = sheet.max_row - 1	# This takes time, so it is stored.

	index = 0
	tableLayout = ""

	for row in sheet.iter_rows(min_row=2, min_col=2,values_only=True):
		index += 1

		state, county, pop = row

		countyData.setdefault(state, {'tracts': 0, 'pop': 0, 'counties': {}})
		countyData[state]['tracts'] += 1
		countyData[state]['pop'] += pop
		countyData[state]['counties'].setdefault(county, {'tracts': 0, 'pop': 0})
		countyData[state]['counties'][county]['tracts'] += 1
		countyData[state]['counties'][county]['pop'] += pop

		attribute = ""
		
		if prevCounty != county:
			attribute = " id=\"{}.{}\"".format(state,county)
			prevCounty = county
			
		tableLayout += tableItem.format(attribute,index+1,state,county,pop)

		if not ( index % 2500 ) or ( index == limit ):
			dom.append_layout('table', tableLayout)
			dom.scroll_to(dom.last_child('table'))
			dom.flush()
			dom.set_content('output', 'Reading rows {}/{}'.format(index,limit))
			tableLayout = ""

	dom.set_content('output', 'Calculating...')
	
	statesLayout = ""
	countiesLayout = ""

	for state, stateData in countyData.items():
		statesLayout += statesItem.format(State=state)

		countiesLayout += stateInCountiesItem.format(State=state,Pop=stateData['pop'],Tracts=stateData['tracts'])

		for county, data in stateData['counties'].items():
			countiesLayout += countiesItem.format(State=state,County=county,Pop=data['pop'],Tracts=data['tracts'])

	dom.set_layout("states", statesLayout)
	dom.set_layout("counties", countiesLayout)
	dom.set_content('output', 'Done')
	dom.add_class("output", "hidden")
	
def ac_connect(dom):
	dom.set_layout("", open("Main.html").read())
	reading(dom)

def ac_view(dom,id):
	dom.scroll_to(dom.get_content(id))

callbacks = {
	"": ac_connect,
	"View": ac_view,
}

Atlas.launch(callbacks, None, open("Head.html").read())
