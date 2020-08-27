# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2020 Claude SIMON (https://q37.info/s/rmnmqd49)

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import os, sys

os.chdir(os.path.dirname(os.path.realpath(__file__)))
sys.path.append("../../atlastk")

# Note to the dev: 'openpyxl' must be imported before 'atlastk',
# as the 'load_workbook' function from 'openpyxl' is overloaded.
# Applies only in DEV context.
import openpyxl
import atlastk as Atlas

tableItem = """
<tr{}>
	<td>{}</td>
	<td>{}</td>
	<td>{}</td>
	<td>{}</td>
	<td>{}</td>
</tr>
"""

statesItem = """
<tr data-xdh-content="{State}" data-xdh-onevent="View" style="cursor: default;" title="Pop: {Pop}, tracts: {Tracts}">
	<td>{State}</td>
	<td>{Pop}</td>
</tr>
"""

countiesItem = """
<tr data-xdh-onevent="View" data-xdh-content="{State}.{County}" style="cursor: default;">
	<td>{County}</td>
	<td>{Pop}</td>
	<td>{Tracts}</td>
</tr>
"""

stateInCountiesItem = """
<tr id=\"{State}\" class="state">
	<td><span style=\"font-style: oblique\">State</span>: {State}</td>
	<td>{Pop}</td>
	<td>{Tracts}</td>
</tr>
"""

countyData = {}

def reading(dom):
	global countyData

	countyData = {}

	dom.set_contents({
		"output": "Initialization…",
		"table": "",
		"counties": "",
		"states": ""
	})

	dom.remove_class("output", "hidden")
	prevCounty = ""

	dom.set_content('output', 'Opening workbook...')
	wb = openpyxl.load_workbook(dom.get_content("set"),read_only=True)

	sheet = wb['Population by Census Tract']

	dom.set_content('output', 'Reading rows...')

	limit = sheet.max_row - 1	# This takes time, so it is stored.

	index = 0
	tableLayout = ""

	for row in sheet.iter_rows(min_row=2,values_only=True):
		index += 1

		tract, state, county, pop = row

		countyData.setdefault(state, {'tracts': 0, 'pop': 0, 'counties': {}})
		countyData[state]['tracts'] += 1
		countyData[state]['pop'] += pop
		countyData[state]['counties'].setdefault(county, {'tracts': 0, 'pop': 0})
		countyData[state]['counties'][county]['tracts'] += 1
		countyData[state]['counties'][county]['pop'] += pop

		attribute = ""
		
		if prevCounty != county:
			attribute = " id=\"{}.{}\"".format(state,county)
			prevCounty = county
			
		tableLayout += tableItem.format(attribute,index+1,tract,state,county,pop)

		if not ( index % 2500 ) or ( index == limit ):
			dom.end('table', tableLayout)
			dom.scroll_to(dom.last_child('table'))
			dom.flush()
			dom.set_content('output', 'Reading rows {}/{}'.format(index,limit))
			tableLayout = ""

	dom.set_content('output', 'Calculating...')
	
	statesLayout = ""
	countiesLayout = ""

	for state, stateData in countyData.items():
		tracts, pop, counties = stateData.values()
		statesLayout += statesItem.format(State=state,Pop=pop,Tracts=tracts)

		countiesLayout += stateInCountiesItem.format(State=state,Pop=pop,Tracts=tracts)

		for county, data in counties.items():
			countiesLayout += countiesItem.format(State=state,County=county,Pop=data['pop'],Tracts=data['tracts'])

	dom.inner("states", statesLayout)
	dom.inner("counties", countiesLayout)
	dom.set_content('output', 'Done')
	dom.add_class("output", "hidden")
	
def ac_connect(dom):
	dom.inner("", open("Main.html").read())
	reading(dom)

def ac_view(dom,id):
	dom.scroll_to(dom.get_content(id))

callbacks = {
	"": ac_connect,
	"View": ac_view,
	"Refresh": lambda dom : reading(dom)
}

Atlas.launch(callbacks, None, open("Head.html").read())
